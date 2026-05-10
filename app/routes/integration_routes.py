from flask import Blueprint, render_template, request, redirect, url_for, jsonify, flash, current_app
from flask_login import login_required, current_user
from app import csrf
from app.services.monobank_service import MonobankService
from app.repositories.user_repository import get_monobank_token, save_monobank_token, delete_monobank_token
from app.repositories.account_repository import get_accounts_by_user, create_account, get_multi_user_accounts, get_account_by_id, update_account_balance
from app.repositories.transaction_repository import add_transaction
from app.models import db, get_current_time
from datetime import datetime, timezone
from app.repositories.partnership_repository import get_active_partnership
from app.repositories.category_repository import get_multi_user_categories, get_categories_by_user, resolve_category_name
from app.config import Config
from app.services.ai_service import AIService
from app.utils.strings import get_string
import random
import threading
import tempfile
import os
import json

integration_bp = Blueprint('integration', __name__)

_processing_status = {}
_processing_results = {}

def _processing_key(user_id):
    return f"csv_processing_{user_id}"

def _results_key(user_id):
    return f"csv_results_{user_id}"

@integration_bp.route('/integrations')
@login_required
def integrations():
    mono_token = get_monobank_token(current_user.id)
    accounts = get_accounts_by_user(current_user.id, is_shared=False)
    partnership = get_active_partnership(current_user.id)
    shared_accounts = []
    if partnership:
        partner_id = partnership.user1_id if partnership.user2_id == current_user.id else partnership.user2_id
        shared_accounts = get_multi_user_accounts([current_user.id, partner_id], is_shared=True)
    
    proc_key = _processing_key(current_user.id)
    processing = _processing_status.get(proc_key, {})
    
    return render_template('integrations.html', 
                         username=current_user.username, 
                         is_mono_connected=bool(mono_token), 
                         accounts=accounts, 
                         shared_accounts=shared_accounts, 
                         has_shared_budget=bool(partnership),
                         processing=processing)

@integration_bp.route('/sync_monobank', methods=['POST'])
@login_required
def sync_monobank():
    form_token = request.form.get('monobank_token')
    db_token_record = get_monobank_token(current_user.id)
    token = form_token or (db_token_record.token if db_token_record else None)

    if not token: 
        return redirect(url_for('integration.integrations'))
    if form_token: 
        save_monobank_token(current_user.id, token)

    client_info = MonobankService.get_client_info(token)
    if client_info and 'accounts' in client_info:
        main_card = client_info['accounts'][0]
        real_balance = main_card.get('balance', 0) / 100.0
        account_name = 'Monobank'

        from app.repositories.account_repository import get_accounts_by_user as get_accs
        mono_account = next((a for a in get_accs(current_user.id) if a.name == account_name), None)
        if not mono_account:
            mono_account = create_account(account_name, real_balance, current_user.id, False)
        else:
            mono_account.balance = real_balance
            db.session.commit()

        statement = MonobankService.get_statement(token)
        from app.models import Transaction
        pending_transactions = []
        for t in statement:
            amount_uah = round(abs(t.get('amount', 0) / 100.0), 2)
            t_type = 'Expense' if t.get('amount', 0) < 0 else 'Income'
            t_desc = t.get('description', 'Monobank')
            t_date = datetime.fromtimestamp(t.get('time'), tz=timezone.utc).replace(tzinfo=None)

            if not Transaction.query.filter_by(user_id=current_user.id, account_id=mono_account.id,
                                                 amount=amount_uah, type=t_type, description=t_desc, date=t_date).first():
                pending_transactions.append({
                    'type': t_type,
                    'amount': amount_uah,
                    'category': 'Other',
                    'description': t_desc,
                    'date': t_date
                })

        existing_cats = get_categories_by_user(current_user.id, is_shared=False)
        ai_category_choices = AIService.choose_existing_categories(pending_transactions, existing_cats)
        for tx_idx, tx in enumerate(pending_transactions):
            cat_name = ai_category_choices.get(tx_idx) or resolve_category_name(
                tx['category'],
                tx['type'],
                existing_cats,
                current_user.id,
                False,
                description=tx['description'],
                color=random.choice(Config.COLORS_PALETTE)
            )
            add_transaction(tx['type'], cat_name, tx['amount'], tx['description'], tx['date'], current_user.id, mono_account.id, False)

    return redirect(url_for('integration.integrations'))

@integration_bp.route('/unlink_monobank', methods=['POST'])
@login_required
def unlink_monobank():
    delete_monobank_token(current_user.id)
    return redirect(url_for('integration.integrations'))

PRIVAT_CATEGORY_MAP = {
    'Супермаркети та продукти': 'Products',
    'Ресторани та кафе': 'Restaurants',
    'Кафе і ресторани': 'Restaurants',
    'Транспорт': 'Transport',
    'Таксі': 'Transport',
    'АЗС': 'Gas',
    'Розваги': 'Entertainment',
    'Медицина': 'Health',
    'Аптеки': 'Health',
    'Одяг та взуття': 'Clothing',
    'Краса': 'Beauty',
    'Зарахування': 'Income',
    'Зарахування переказу': 'Income',
    'Переказ': 'Transfer',
    'Дім та ремонт': 'Home',
    'Комунальні послуги': 'Utilities',
    'Заощадження': 'Savings',
    'Скарбничка': 'Savings',
    'Телекомунікації': 'Communication',
    'Освіта': 'Education',
    'Подорожі': 'Travel',
    'Готелі': 'Hotel',
    'Спорт': 'Sport',
    'Фінансові послуги': 'Bank',
    'Штрафи та збори': 'Fines',
}

def parse_generic_xlsx(file_bytes):
    """Parse ANY XLSX file — remove empty columns, return compact text + raw rows."""
    try:
        import openpyxl, io
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        raw_rows = []  
        row_count = 0

        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                row_data = []
                for cell in row:
                    if cell is None:
                        row_data.append('')
                    else:
                        cell_str = str(cell).strip()
                        if len(cell_str) > 30:
                            cell_str = cell_str[:30]
                        row_data.append(cell_str)
                raw_rows.append(row_data)
                row_count += 1
                if row_count >= 1000:
                    break
            if row_count >= 1000:
                break

        if not raw_rows:
            return "", []

        col_count = max(len(r) for r in raw_rows)
        non_empty_cols = []
        for col_idx in range(col_count):
            has_value = False
            for row in raw_rows[1:]:   
                val = row[col_idx] if col_idx < len(row) else ''
                if val and val.strip():
                    has_value = True
                    break
            if has_value:
                non_empty_cols.append(col_idx)

        compact_rows = []
        for row in raw_rows:
            filtered = [row[i] if i < len(row) else '' for i in non_empty_cols]
            compact_rows.append('\t'.join(filtered))

        text = '\n'.join(compact_rows)
        print(f"XLSX parsed: {row_count} rows, {len(non_empty_cols)} active cols, {len(text)} chars")
        return text, raw_rows
    except Exception as e:
        print(f"XLSX parse error: {e}")
        import traceback; traceback.print_exc()
        return "", []


def parse_privat_xlsx(file_bytes):
    """Parse PrivatBank XLSX — row 2 is header, row 3+ are data.
    Columns: A=date, B=bank_category, C=card, D=description, E=amount, I=balance
    """
    import openpyxl, io, re
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    transactions = []
    end_balance = None

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx <= 2:
            continue
        if not row or row[0] is None or row[4] is None:
            continue

        date_str = str(row[0]).strip()   
        category_raw = str(row[1]).strip() if row[1] else ''
        description = str(row[3]).strip() if row[3] else category_raw
        amount_raw = row[4]              
        balance_raw = row[8] if len(row) > 8 else None  

        if end_balance is None and balance_raw is not None:
            try:
                end_balance = float(str(balance_raw).replace(' ', '').replace(',', '.'))
            except Exception:
                pass

        try:
            dt = datetime.strptime(date_str[:10], '%d.%m.%Y')
            date_iso = dt.strftime('%Y-%m-%d')
        except Exception:
            continue

        try:
            amount = float(str(amount_raw).replace(' ', '').replace(',', '.'))
        except Exception:
            continue

        if amount == 0:
            continue

        t_type = 'Income' if amount > 0 else 'Expense'
        INCOME_CATEGORIES = {'Зарахування', 'Зарахування переказу', 'Кешбек', 'Зарахування кешбеку'}
        if category_raw in INCOME_CATEGORIES:
            t_type = 'Income'
        amount_abs = abs(amount)
        category_hint = PRIVAT_CATEGORY_MAP.get(category_raw, category_raw or 'Other')

        transactions.append({
            'date': date_iso,
            'description': description[:100],
            'category': category_hint,
            'type': t_type,
            'amount': round(amount_abs, 2),
        })

    print(f"PrivatBank parsed: {len(transactions)} transactions, balance={end_balance}")
    return transactions, end_balance


def parse_abank_xlsx(file_bytes):
    """Parse A-Bank XLSX — rows 1-19 are meta, row 20 is header, row 21+ are data.
    Columns: A=datetime, B=card, C=description, D=MCC, E=amount(UAH), K=balance_after
    Balance at end of period is in row 15, column H.
    """
    import openpyxl, io, re
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    transactions = []
    end_balance = None

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx == 15:
            cell_h = row[7] if len(row) > 7 else None
            if cell_h:
                m = re.search(r'([\d\s]+[\.,]\d{2})', str(cell_h))
                if m:
                    try:
                        end_balance = float(m.group(1).replace(' ', '').replace(',', '.'))
                    except Exception:
                        pass
        if row_idx >= 20:
            break

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx <= 20:   
            continue
        if not row or row[0] is None or row[4] is None:
            continue

        date_str = str(row[0]).strip()   
        description = str(row[2]).strip() if row[2] else ''  
        amount_raw = row[4]              

        try:
            dt = datetime.strptime(date_str[:10], '%d.%m.%Y')
            date_iso = dt.strftime('%Y-%m-%d')
        except Exception:
            continue

        try:
            amount = float(str(amount_raw).replace(' ', '').replace(',', '.'))
        except Exception:
            continue

        if amount == 0:
            continue

        t_type = 'Income' if amount > 0 else 'Expense'
        amount_abs = abs(amount)

        transactions.append({
            'date': date_iso,
            'description': description[:100],
            'category': '',   
            'type': t_type,
            'amount': round(amount_abs, 2),
        })

    print(f"A-Bank parsed: {len(transactions)} transactions, balance={end_balance}")
    return transactions, end_balance


def parse_csv_generic(text, existing_categories=None, is_xlsx=False):
    """Use AI to parse any CSV/XLSX format."""
    return AIService.parse_statement(text, existing_categories, is_xlsx=is_xlsx)


def _process_bank_background(app, user_id, transactions, end_balance, account_id, is_shared, existing_cats, lang):
    """Background thread: data already parsed by code, AI only assigns categories."""
    proc_key = _processing_key(user_id)
    results_key = _results_key(user_id)
    with app.app_context():  
        try:
            bank_src = _processing_status.get(proc_key, {}).get('bank_source', 'privat')
            _processing_status[proc_key] = {
                'status': 'processing',
                'message': 'Підбираємо категорії через ШІ...',
                'progress': 40,
                'bank_source': bank_src,
            }

            if not transactions:
                _processing_status[proc_key] = {
                    'status': 'completed',
                    'message': get_string('csv_status_no_tx', lang=lang),
                    'progress': 100,
                    'count': 0,
                    'bank_source': bank_src,
                }
                return

            _processing_status[proc_key]['message'] = f'Категоризуємо {len(transactions)} транзакцій...'
            _processing_status[proc_key]['progress'] = 60

            ai_category_choices = AIService.choose_existing_categories(transactions, existing_cats)

            _processing_status[proc_key]['progress'] = 85
            normalized = []
            for tx_idx, tx in enumerate(transactions):
                t_type = tx.get('type', 'Expense')
                description = tx.get('description', '')
                ai_cat = ai_category_choices.get(tx_idx)
                tx['category'] = ai_cat or resolve_category_name(
                    tx.get('category', 'Other'),
                    t_type,
                    existing_cats,
                    user_id,
                    is_shared,
                    description=description,
                    color=random.choice(Config.COLORS_PALETTE)
                )
                normalized.append(tx)

            preview_data = {
                'account_id': account_id,
                'transactions': normalized,
                'end_balance': end_balance,
                'is_shared': is_shared
            }
            with open(_preview_path(user_id), 'w', encoding='utf-8') as f:
                json.dump(preview_data, f, ensure_ascii=False)

            _processing_results[results_key] = preview_data
            _processing_status[proc_key] = {
                'status': 'completed',
                'message': f'Оброблено {len(normalized)} транзакцій',
                'progress': 100,
                'count': len(normalized),
                'bank_source': bank_src,
            }

        except Exception as e:
            print(f'Bank background error: {e}')
            import traceback; traceback.print_exc()
            _processing_status[proc_key] = {
                'status': 'error',
                'message': f'Помилка обробки: {str(e)}',
                'progress': 100,
                'bank_source': bank_src,
            }

def _preview_path(user_id):
    """Path to temp file that stores preview data for this user."""
    tmp = tempfile.gettempdir()
    return os.path.join(tmp, f'finapp_preview_{user_id}.json')

def _process_csv_background(user_id, file_bytes, filename, account_id, is_shared, valid_user_ids, existing_cats, lang):
    """Background thread function to process CSV/XLSX file."""
    proc_key = _processing_key(user_id)
    results_key = _results_key(user_id)
    
    try:
        _processing_status[proc_key] = {
            'status': 'processing',
            'message': get_string('ai_processing', lang=lang),
            'progress': 20,
            'bank_source': 'csv',
        }
        
        is_xlsx = filename.endswith('.xlsx')
        end_balance = None
        transactions = []
        
        if is_xlsx:
            _processing_status[proc_key]['message'] = 'Extracting data from XLSX...'
            _processing_status[proc_key]['progress'] = 30
            text, _raw_rows = parse_generic_xlsx(file_bytes)
            
            import re
            balance_patterns = [
                r'кінцевий залишок[\s:]*([\d\s\.,]+)',
                r'balance[\s:]*([\d\s\.,]+)',
                r'залишок[\s:]*([\d\s\.,]+)'
            ]
            for pattern in balance_patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    try:
                        balance_str = match.group(1).strip().replace(' ', '').replace(',', '.')
                        end_balance = float(balance_str)
                        break
                    except:
                        pass
            
            _processing_status[proc_key]['message'] = 'Analyzing with AI...'
            _processing_status[proc_key]['progress'] = 50
            transactions, ai_balance = parse_csv_generic(text, existing_cats, is_xlsx=True)
            if ai_balance is not None:
                end_balance = ai_balance
        else:
            _processing_status[proc_key]['message'] = 'Reading CSV text...'
            _processing_status[proc_key]['progress'] = 30
            try:
                text = file_bytes.decode('utf-8')
            except UnicodeDecodeError:
                text = file_bytes.decode('cp1251', errors='replace')
            lines = text.splitlines()
            truncated = '\n'.join(lines[:500])
            
            _processing_status[proc_key]['message'] = 'Analyzing with AI...'
            _processing_status[proc_key]['progress'] = 50
            transactions, ai_balance = parse_csv_generic(truncated, existing_cats, is_xlsx=False)
            if ai_balance is not None:
                end_balance = ai_balance

        if transactions is None:
            _processing_status[proc_key] = {
                'status': 'error',
                'message': get_string('csv_status_ai_failed', lang=lang),
                'progress': 100,
        'bank_source': 'csv',
    }
            return

        if not transactions:
            _processing_status[proc_key] = {
                'status': 'completed',
                'message': get_string('csv_status_no_tx', lang=lang),
                'progress': 100,
                'count': 0,
            'bank_source': 'csv',
        }
            return
        
        _processing_status[proc_key]['message'] = 'Categorizing transactions...'
        _processing_status[proc_key]['progress'] = 80
        ai_category_choices = AIService.choose_existing_categories(transactions, existing_cats)
        normalized_transactions = []
        for tx_idx, tx in enumerate(transactions):
            t_type = tx.get('type', 'Expense')
            description = tx.get('description', '')
            tx['category'] = ai_category_choices.get(tx_idx) or resolve_category_name(
                tx.get('category', 'Other'),
                t_type,
                existing_cats,
                user_id,
                is_shared,
                description=description,
                color=random.choice(Config.COLORS_PALETTE)
            )
            normalized_transactions.append(tx)
        
        preview_data = {
            'account_id': account_id,
            'transactions': normalized_transactions,
            'end_balance': end_balance,
            'is_shared': is_shared
        }
        
        with open(_preview_path(user_id), 'w', encoding='utf-8') as f:
            json.dump(preview_data, f, ensure_ascii=False)
        
        _processing_results[results_key] = preview_data
        _processing_status[proc_key] = {
            'status': 'completed',
            'message': f'Successfully processed {len(normalized_transactions)} transactions',
            'progress': 100,
            'count': len(normalized_transactions),
            'bank_source': 'csv',
        }
        
    except Exception as e:
        print(f'Background processing error: {e}')
        import traceback; traceback.print_exc()
        _processing_status[proc_key] = {
            'status': 'error',
            'message': f'Processing error: {str(e)}',
            'progress': 100,
            'bank_source': 'csv',
        }

@integration_bp.route('/upload_csv', methods=['POST'])
@login_required
def upload_csv():
    if 'csv_file' not in request.files:
        flash('File not uploaded', 'error')
        return redirect(url_for('integration.integrations'))

    file = request.files['csv_file']
    if not file or file.filename == '':
        flash('File not uploaded', 'error')
        return redirect(url_for('integration.integrations'))

    account_id = request.form.get('account_id')
    if not account_id:
        flash('Select account', 'error')
        return redirect(url_for('integration.integrations'))

    acc = get_account_by_id(int(account_id))
    partnership = get_active_partnership(current_user.id)
    partner_id = partnership.user1_id if partnership and partnership.user2_id == current_user.id else partnership.user2_id if partnership else None
    valid_user_ids = [current_user.id] + ([partner_id] if partner_id else [])
    if not acc or acc.user_id not in valid_user_ids or (acc.is_shared and not partnership):
        flash('Account not found', 'error')
        return redirect(url_for('integration.integrations'))

    existing_cats = get_multi_user_categories(valid_user_ids, is_shared=True) if acc.is_shared else get_categories_by_user(current_user.id, is_shared=False)

    try:
        filename = file.filename.lower()
        file_bytes = file.read()
        
        proc_key = _processing_key(current_user.id)
        _processing_status[proc_key] = {
            'status': 'starting',
            'message': 'Starting processing...',
            'progress': 10,
            'bank_source': 'csv'
        }
        
        from app.utils.strings import get_current_lang
        lang = get_current_lang()
        
        thread = threading.Thread(
            target=_process_csv_background,
            args=(current_user.id, file_bytes, filename, account_id, acc.is_shared, valid_user_ids, existing_cats, lang)
        )
        thread.daemon = True
        thread.start()
        
        flash(get_string('csv_processing_desc'), 'info')
        return redirect(url_for('integration.integrations'))
        
    except Exception as e:
        print(f'Upload error: {e}')
        import traceback; traceback.print_exc()
        flash('File upload error. Try again.', 'error')
        return redirect(url_for('integration.integrations'))

# ─── PrivatBank 
@integration_bp.route('/upload_privat', methods=['POST'])
@login_required
def upload_privat():
    """Upload a PrivatBank XLSX statement — parsed by code, AI categorizes only."""
    if 'privat_file' not in request.files:
        flash('Файл не завантажено', 'error')
        return redirect(url_for('integration.integrations'))
    file = request.files['privat_file']
    if not file or file.filename == '':
        flash('Файл не завантажено', 'error')
        return redirect(url_for('integration.integrations'))
    account_id = request.form.get('privat_account_id')
    if not account_id:
        flash('Оберіть рахунок', 'error')
        return redirect(url_for('integration.integrations'))

    acc = get_account_by_id(int(account_id))
    partnership = get_active_partnership(current_user.id)
    partner_id = partnership.user1_id if partnership and partnership.user2_id == current_user.id else partnership.user2_id if partnership else None
    valid_user_ids = [current_user.id] + ([partner_id] if partner_id else [])
    if not acc or acc.user_id not in valid_user_ids or (acc.is_shared and not partnership):
        flash('Рахунок не знайдено', 'error')
        return redirect(url_for('integration.integrations'))

    existing_cats = get_multi_user_categories(valid_user_ids, is_shared=True) if acc.is_shared else get_categories_by_user(current_user.id, is_shared=False)

    try:
        file_bytes = file.read()
        transactions, end_balance = parse_privat_xlsx(file_bytes)

        proc_key = _processing_key(current_user.id)
        _processing_status[proc_key] = {'status': 'starting', 'message': 'Зчитуємо файл ПриватБанку...', 'progress': 10, 'bank_source': 'privat'}

        from app.utils.strings import get_current_lang
        lang = get_current_lang()

        thread = threading.Thread(
            target=_process_bank_background,
            args=(current_app._get_current_object(), current_user.id, transactions, end_balance, account_id, acc.is_shared, existing_cats, lang)
        )
        thread.daemon = True
        thread.start()

        flash('Файл ПриватБанку прийнято, обробляємо...', 'info')
        return redirect(url_for('integration.integrations'))
    except Exception as e:
        print(f'PrivatBank upload error: {e}')
        import traceback; traceback.print_exc()
        flash('Помилка завантаження файлу ПриватБанку.', 'error')
        return redirect(url_for('integration.integrations'))


# ─── A-Bank upload 
@integration_bp.route('/upload_abank', methods=['POST'])
@login_required
def upload_abank():
    """Upload an A-Bank XLSX statement — parsed by code, AI categorizes only."""
    if 'abank_file' not in request.files:
        flash('Файл не завантажено', 'error')
        return redirect(url_for('integration.integrations'))
    file = request.files['abank_file']
    if not file or file.filename == '':
        flash('Файл не завантажено', 'error')
        return redirect(url_for('integration.integrations'))
    account_id = request.form.get('abank_account_id')
    if not account_id:
        flash('Оберіть рахунок', 'error')
        return redirect(url_for('integration.integrations'))

    acc = get_account_by_id(int(account_id))
    partnership = get_active_partnership(current_user.id)
    partner_id = partnership.user1_id if partnership and partnership.user2_id == current_user.id else partnership.user2_id if partnership else None
    valid_user_ids = [current_user.id] + ([partner_id] if partner_id else [])
    if not acc or acc.user_id not in valid_user_ids or (acc.is_shared and not partnership):
        flash('Рахунок не знайдено', 'error')
        return redirect(url_for('integration.integrations'))

    existing_cats = get_multi_user_categories(valid_user_ids, is_shared=True) if acc.is_shared else get_categories_by_user(current_user.id, is_shared=False)

    try:
        file_bytes = file.read()
        transactions, end_balance = parse_abank_xlsx(file_bytes)

        proc_key = _processing_key(current_user.id)
        _processing_status[proc_key] = {'status': 'starting', 'message': 'Зчитуємо файл А-Банку...', 'progress': 10, 'bank_source': 'abank'}

        from app.utils.strings import get_current_lang
        lang = get_current_lang()

        thread = threading.Thread(
            target=_process_bank_background,
            args=(current_app._get_current_object(), current_user.id, transactions, end_balance, account_id, acc.is_shared, existing_cats, lang)
        )
        thread.daemon = True
        thread.start()

        flash('Файл А-Банку прийнято, обробляємо...', 'info')
        return redirect(url_for('integration.integrations'))
    except Exception as e:
        print(f'A-Bank upload error: {e}')
        import traceback; traceback.print_exc()
        flash('Помилка завантаження файлу А-Банку.', 'error')
        return redirect(url_for('integration.integrations'))


@integration_bp.route('/csv_status')
@login_required
def csv_status():
    """Return processing status as JSON."""
    proc_key = _processing_key(current_user.id)
    status = _processing_status.get(proc_key, {'status': 'none'})
    return jsonify(status)

@integration_bp.route('/csv_cancel', methods=['POST'])
@login_required
def csv_cancel():
    """Cancel CSV processing — clear status AND delete preview file."""
    proc_key = _processing_key(current_user.id)
    if proc_key in _processing_status:
        del _processing_status[proc_key]
    preview_file = _preview_path(current_user.id)
    try:
        if os.path.exists(preview_file):
            os.remove(preview_file)
    except Exception:
        pass
    flash(get_string('csv_processing_cancelled'), 'info')
    return redirect(url_for('integration.integrations'))

@integration_bp.route('/csv_preview', methods=['GET', 'POST'])
@login_required
@csrf.exempt
def csv_preview():
    path = _preview_path(current_user.id)
    
    if request.method == 'POST':
        selected_indices = request.form.getlist('selected')
        try:
            with open(path, encoding='utf-8') as f:
                data = json.load(f)
        except Exception:
            flash(get_string('csv_error_no_data'), 'error')
            return redirect(url_for('integration.integrations'))
        
        acc = get_account_by_id(int(data['account_id']))
        all_transactions = data['transactions']
        is_shared = bool(data.get('is_shared'))
        
        added = 0
        update_balance_flag = request.form.get('update_balance') == 'yes' and data.get('end_balance') is not None
        
        for idx_str in selected_indices:
            try:
                pt = all_transactions[int(idx_str)]
                date_obj = datetime.strptime(pt['date'], '%Y-%m-%d')
                amt = float(pt['amount'])
                user_cat = request.form.get(f'category_{idx_str}')
                final_cat = user_cat.strip() if user_cat and user_cat.strip() else pt['category']
                raw_type = pt['type']
                if raw_type == 'Income':
                    normalized_type = 'Дохід'
                elif raw_type == 'Expense':
                    normalized_type = 'Витрата'
                else:
                    normalized_type = raw_type  
                add_transaction(
                    t_type=normalized_type,
                    category=final_cat,
                    amount=amt,
                    description=pt.get('description', 'Import'),
                    date=date_obj,
                    user_id=current_user.id,
                    account_id=acc.id,
                    is_shared=is_shared
                )
                if not update_balance_flag:
                    update_account_balance(acc, amt, normalized_type)
                added += 1
            except Exception as e:
                print(f'Import row error: {e}')
        
        if update_balance_flag:
            acc.balance = float(data['end_balance'])
        
        db.session.commit()
        try:
            os.remove(path)
            proc_key = _processing_key(current_user.id)
            if proc_key in _processing_status:
                del _processing_status[proc_key]
        except Exception:
            pass
        
        flash(get_string('success_csv_import', count=added), 'success')
        return redirect(url_for('shared.shared_budget' if is_shared else 'budget.home'))
    
    proc_key = _processing_key(current_user.id)
    status = _processing_status.get(proc_key, {})
    
    if status.get('status') != 'completed':
        return render_template('csv_processing.html', 
                             status=status,
                             username=current_user.username)
    
    try:
        with open(path, encoding='utf-8') as f:
            data = json.load(f)
    except Exception:
        flash(get_string('csv_error_no_data'), 'error')
        return redirect(url_for('integration.integrations'))
    
    acc = get_account_by_id(int(data['account_id']))
    is_shared = bool(data.get('is_shared'))
    transactions = data['transactions']
    end_balance = data.get('end_balance')

    if is_shared:
        partnership = get_active_partnership(current_user.id)
        if partnership:
            valid_user_ids = list({partnership.user1_id, partnership.user2_id})
        else:
            valid_user_ids = [current_user.id]
        categories = get_multi_user_categories(valid_user_ids, is_shared=True)
    else:
        categories = get_categories_by_user(current_user.id, is_shared=False)

    from flask import make_response
    response = make_response(render_template('csv_preview.html',
                           transactions=transactions,
                           account=acc,
                           end_balance=end_balance,
                           categories=categories,
                           username=current_user.username))
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response
